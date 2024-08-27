import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter


def create_excel_mapping(data: list, filename: str) -> None:
    # create new book excel
    wb = openpyxl.Workbook()

    # Create new execel list
    ws = wb.active
    ws.append(["start"])

    # Create headers
    source_target: list[str] = ["#", "Тип объекта"]
    source: list[str] = ["Система", "Топик Kafka 610_4", "База данных","Таблица", "Описание таблицы", "Имя поля","Комментарий поля", "Тип данных", "PK", "Not Null"]
    target: list[str] = ["База/Система", "Схема", "Таблица", "Название родительской таблицы", "Описание таблицы", "Код атрибута", "Описание атрибута", "Комментарий", "Тип данных", "Length", "PK", "FK", "Not Null", "Rejectable", "Trace New Values"]
    len_sourcetarget: int = len(source_target)
    len_source: int = len(source)
    len_target: int = len(target)
    headers: list[str] = source_target + source + target

    ws.column_dimensions['A'].width = len_sourcetarget + 1
    ws.column_dimensions['B'].width = len_source + 1
    ws.column_dimensions['C'].width = len_target + 1

    ws.append(headers)
    for item in data:
        ws.append(item)

    for col in ws.columns:
        max_length: int = 0
        column: list = col[1:]
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length: int = len(str(cell.value))

        adjusted_width: int = (max_length + 2)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

    fiolet_fill = PatternFill(start_color='B2AFE0', end_color='B2AFE0', fill_type='solid')
    orange_fill = PatternFill(start_color='ffd966', end_color='ffd966', fill_type='solid')
    blue_fill = PatternFill(start_color='A5C4E0', end_color='A5C4E0', fill_type='solid')
    green_fill = PatternFill(start_color='aed59d', end_color='aed59d', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sourcetarget_letter = get_column_letter(len_sourcetarget)
    start_source_letter = get_column_letter(len_sourcetarget + 1)
    end_source_letter = get_column_letter(len_source+len_sourcetarget)
    start_target_letter = get_column_letter(len_sourcetarget+len_source + 1)
    end_target_letter = get_column_letter(len_sourcetarget+len_source+len_target)
    for col in range(1, len_sourcetarget + 1):
        for cell in ws[f"A{col}":f"{sourcetarget_letter}{col}"][0]:
            cell.fill = fiolet_fill
        for cell in ws[f"{start_target_letter}{col}":f"{end_target_letter}{col}"][0]:
            cell.fill = blue_fill

        for cell in ws[f"A{col}:{end_target_letter}{col}"][0]:
            cell.border = border

    for cell in ws[f"{start_source_letter}{1}":f"{end_source_letter}{1}"][0]:
        cell.fill = orange_fill
    for cell in ws[f"{start_source_letter}{2}":f"{end_source_letter}{2}"][0]:
        cell.fill = green_fill


    ws.merge_cells(f"A1:{sourcetarget_letter}1")
    ws["A1"] = "Source/Target"

    ws.merge_cells(f"{start_source_letter}1:{end_source_letter}1")
    ws[f"{start_source_letter}1"] = "Source"

    ws.merge_cells(f"{start_target_letter}1:{end_target_letter}1")
    ws[f"{start_target_letter}1"] = "Target"

    # Выравнивание текста по центру в объединенных ячейках
    center_alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = center_alignment
    ws[f'{start_source_letter}1'].alignment = center_alignment
    ws[f'{start_target_letter}1'].alignment = center_alignment

    wb.save(f"{filename}")