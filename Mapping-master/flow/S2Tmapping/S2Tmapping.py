import re
import os
import typing
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter

if typing.TYPE_CHECKING:
    from flow.flow import Flow

def create_mapping(flow: "Flow", json_file: str, subo_name: str, mapping_version: str, database: str) -> None:
    replace_version: str = json_file.split('_')[-1][2:-5] if len(json_file.split('_')[-1][2:-5]) > 0 else '1.0'
    file_name: str = os.path.join(os.path.dirname(json_file), f'S2T_mapping_{subo_name}_'
                             + os.path.basename(json_file).replace('json', 'xlsx')).replace(replace_version, str(mapping_version))

    # create new book excel
    wb = openpyxl.Workbook()

    # Create new execel list
    ws = wb.active
    ws.append(["start"])

    # Create headers
    source_target: list[str] = ["#", "Тип объекта"]
    source: list[str] = ["База/Система", "Класс", "Наименование класса","Тэг в JSON", "Описание Тэга", "Тип данных","Длина", "PK", "FK", "Not Null"]
    target: list[str] = ["База/Система", "Схема", "Таблица", "Название родительской таблицы", "Описание таблицы", "Код атрибута", "Описание атрибута", "Комментарий", "Тип данных", "Length", "PK", "FK", "Not Null", "Rejectable", "Trace New Values"]
    len_sourcetarget: int = len(source_target)
    len_source: int = len(source)
    len_target: int = len(target)
    headers: list[str] = source_target + source + target

    ws.column_dimensions['A'].width = len_sourcetarget + 1
    ws.column_dimensions['B'].width = len_source + 1
    ws.column_dimensions['C'].width = len_target + 1

    ws.append(headers)

    index: int = 1
    for values in flow.transform.new_flow.tables:
        source_table: str = re.sub(r'^.*?_', '', values.table_name)
        schema: str = f"prod_repl_subo_{database}"
        for column_data in values.attributes.parsedColumns:

            # Will added manually
            # if "array" in column_data.name:
            #     tag_name = column_data.name.replace("array", "hash")
            # elif "." not in column_data.name:
            #     tag_name = column_data.name
            # else:
            #     tag_name = ".".join(column_data.name.split(".")[1:]) if column_data.colType != "hash" else ".".join(
            #         column_data.name.split(".")[1:]) + "_hash"

            notnull: str = "+" if column_data.name.lower() in ["changeid", "changetype", "hdp_processed_dttm"] else ""
            if column_data.name.lower() in ["changeid", "changetype", "changetimestamp"]:
                comment: str = "Техническое поле"
                tag_json: str = column_data.name
                tag_descr: str = column_data.description
                tag_colType: str = ""
                attr_colType: str = column_data.colType
            elif column_data.name == "hdp_processed_dttm":
                comment: str = "Техническое поле"
                tag_json: str = ""
                tag_descr: str = ""
                tag_colType: str = ""
                attr_colType: str = column_data.colType
            elif "hash" in column_data.colType:
                comment: str = column_data.comment
                tag_json: str = ""
                tag_descr: str = ""
                tag_colType: str = ""
                attr_colType: str = "string"
            else:
                comment: str = ""
                tag_descr: str = f"{values.describe_table}. {column_data.description}" if column_data.description else f"{values.describe_table}"
                tag_colType: str = column_data.colType
                # attr_colType: str = "string" 
                attr_colType: str = 'decimal(38,2)' if column_data.colType in ('number', 'decimal') else ('bigint' if column_data.colType in ('integer', 'bigserial') else 'string')
                if len(values.attributes.explodedColumns) == 1:
                    tag_json: str = ".".join(column_data.name.split(".")[1:])
                else:
                    tag_path: str = '.'.join(column_data.name.split('.')[1:])
                    arr_name: str = column_data.name.split('.')[0]
                    tag_json: str = f"{arr_name}[].{tag_path}" if len(tag_path) > 0 else f"{arr_name}[]"

            if column_data.alias is not None:
                code_attr: str = column_data.alias
            else:
                code_attr: str = column_data.name


            row_data: list = [
                        index,                      # "#"
                        "Реплика",                  # "Тип объекта"
                        subo_name,                  # "База/Система"
                        flow.extract.meta_class,    # "Класс"
                        "",                         # "Наименование класса"
                        tag_json,                   # "Тэг в JSON"
                        tag_descr,                  # "Описание Тэга"
                        tag_colType,                 # "Тип данных"
                        "",                         # "Длина"
                        "",                         # "PK"
                        "",                         # "FK"
                        "",                         # "Not Null"
                        "1642_19 Озеро данных",     # "База/Система"
                        schema,                     # "Схема"
                        values.table_name,          # "Таблица"
                        values.parent_table,        # "Название родительской таблицы"
                        values.describe_table,      # "Описание таблицы"
                        code_attr,                  # "Код атрибута"
                        column_data.description,    # "Описание атрибута"
                        comment,                    # "Комментарий"
                        attr_colType,               # "Тип данных"
                        "",                         # "Length"
                        "",                         # "PK"
                        "",                         # "FK"
                        notnull,                    # "Not Null",
                        "",                         # "Rejectable"
                        "",                         # "Trace New Values"
            ]
            index += 1
            ws.append(row_data)

    for col in ws.columns:
        max_length: int = 0
        column: list = col[1:]
        for cell in column:
            if len(str(cell.value)) > max_length:
                max_length: int = len(str(cell.value))

        adjusted_width: int = (max_length + 2)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

    fiolet_fill = PatternFill(start_color='B2AFE0', end_color='B2AFE0', fill_type='solid')
    orange_fill = PatternFill(start_color='ffd966', end_color='ffd966', fill_type='solid')
    blue_fill = PatternFill(start_color='A5C4E0', end_color='A5C4E0', fill_type='solid')
    green_fill = PatternFill(start_color='aed59d', end_color='aed59d', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sourcetarget_letter = get_column_letter(len_sourcetarget)
    start_source_letter = get_column_letter(len_sourcetarget + 1)
    end_source_letter = get_column_letter(len_source+len_sourcetarget)
    start_target_letter = get_column_letter(len_sourcetarget+len_source + 1)
    end_target_letter = get_column_letter(len_sourcetarget+len_source+len_target)
    for col in range(1, len_sourcetarget + 1):
        for cell in ws[f"A{col}":f"{sourcetarget_letter}{col}"][0]:
            cell.fill = fiolet_fill
        for cell in ws[f"{start_target_letter}{col}":f"{end_target_letter}{col}"][0]:
            cell.fill = blue_fill

        for cell in ws[f"A{col}:{end_target_letter}{col}"][0]:
            cell.border = border

    for cell in ws[f"{start_source_letter}{1}":f"{end_source_letter}{1}"][0]:
        cell.fill = orange_fill
    for cell in ws[f"{start_source_letter}{2}":f"{end_source_letter}{2}"][0]:
        cell.fill = green_fill


    ws.merge_cells(f"A1:{sourcetarget_letter}1")
    ws["A1"] = "Source/Target"

    ws.merge_cells(f"{start_source_letter}1:{end_source_letter}1")
    ws[f"{start_source_letter}1"] = "Source"

    ws.merge_cells(f"{start_target_letter}1:{end_target_letter}1")
    ws[f"{start_target_letter}1"] = "Target"

    # Выравнивание текста по центру в объединенных ячейках
    center_alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = center_alignment
    ws[f'{start_source_letter}1'].alignment = center_alignment
    ws[f'{start_target_letter}1'].alignment = center_alignment

    wb.save(f"{file_name}")