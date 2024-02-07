import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.cell import get_column_letter
import numpy as np
import os

from tkinter import *
from tkinter import messagebox, filedialog

def do_rework(old_mappings_path, new_mappings_path, result_mappings_path):
    old_mappings_files = os.listdir(old_mappings_path)
    for old_file in old_mappings_files:
        if old_file.endswith('.xlsx'):
            try:
                rework_one_mapping(old_mappings_path=old_mappings_path+'/'+old_file, 
                                   new_mappings_path=new_mappings_path+'/'+old_file, 
                                   result_mappings_path=result_mappings_path+'/result_'+old_file)
            except ValueError:
                print("Не найден маппинг " + ValueError)

def rework_one_mapping(old_mappings_path, new_mappings_path, result_mappings_path):
    old_mappings = pd.read_excel(old_mappings_path, sheet_name='Mapping', skiprows=1, index_col=0)
    new_mappings = pd.read_excel(new_mappings_path, sheet_name='Sheet', skiprows=1, index_col=0)


    table_col = 'Таблица.1' if 'Таблица.1' in list(old_mappings.columns.values) else 'Таблица'
    code_col = 'Код атрибута.1' if 'Код атрибута.1' in list(old_mappings.columns.values) else 'Код атрибута'
    type_col = 'Тип данных.1' if 'Тип данных.1' in list(old_mappings.columns.values) else 'Тип данных'
    table_col_new = 'Таблица.1' if 'Таблица.1' in list(new_mappings.columns.values) else 'Таблица'
    code_col_new = 'Код атрибута.1' if 'Код атрибута.1' in list(new_mappings.columns.values) else 'Код атрибута'
    type_col_new = 'Тип данных.1' if 'Тип данных.1' in list(new_mappings.columns.values) else 'Тип данных'

    new_mappings = new_mappings.rename(columns={table_col_new: table_col, code_col_new: code_col, type_col_new: type_col})

    old_mappings_divided = []
    for table in old_mappings[table_col].unique():
        old_mappings_divided.append(old_mappings[old_mappings[table_col] == table])


    new_mappings_divided = []
    for table in new_mappings[table_col].unique():
        new_mappings_divided.append(new_mappings[new_mappings[table_col] == table])

    new_tables = np.setdiff1d(new_mappings[table_col].unique(), old_mappings[table_col].unique())

    result_mappings = pd.DataFrame()

    for old_table in old_mappings_divided:
        is_overlaped = False
        for new_table in new_mappings_divided:
            if new_table.iloc[0][table_col] == old_table.iloc[0][table_col]:
                result_mappings = result_mappings._append(pd.concat([old_table, new_table], ignore_index=True))
                is_overlaped = True
        if not is_overlaped:
            result_mappings = pd.concat([result_mappings, old_table], ignore_index=True)

    for table in new_mappings_divided:
        if table.iloc[0][table_col] in new_tables:
            result_mappings = pd.concat([result_mappings, table], ignore_index=True)

    result_mappings.drop_duplicates(subset=[table_col, code_col, type_col], inplace=True, ignore_index=True)
    result_mappings.index += 1

    wb = openpyxl.Workbook()
    # Create new execel list
    ws = wb.active
    ws.append(["start"])

    # Create headers
    source_target: list[str] = ["#", "Тип объекта"]
    source: list[str] = list(old_mappings.columns.values)[:list(old_mappings.columns.values).index('База/Система.1')]
    target: list[str] = ["База/Система", "Схема", "Таблица", "Название родительской таблицы", "Описание таблицы", "Код атрибута", "Описание атрибута", "Комментарий", "Тип данных", "Length", "PK", "FK", "Not Null", "Rejectable", "Trace New Values"]
    len_sourcetarget: int = len(source_target)
    len_source: int = len(source)
    len_target: int = len(target)
    headers: list[str] = source_target + source + target

    ws.column_dimensions['A'].width = len_sourcetarget + 1
    ws.column_dimensions['B'].width = len_source + 1
    ws.column_dimensions['C'].width = len_target + 1

    ws.append(headers)

    for r in dataframe_to_rows(result_mappings, index=True, header=False):
        ws.append(r)

    ws.delete_rows(3, 1)

    for col in ws.columns:
            max_length: int = 0
            column: list = col[1:]
            for cell in column:
                if len(str(cell.value)) > max_length:
                    max_length: int = len(str(cell.value))

            adjusted_width: int = (max_length + 2)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = adjusted_width

    fiolet_fill = PatternFill(start_color='B2AFE0', end_color='B2AFE0', fill_type='solid')
    orange_fill = PatternFill(start_color='ffd966', end_color='ffd966', fill_type='solid')
    blue_fill = PatternFill(start_color='A5C4E0', end_color='A5C4E0', fill_type='solid')
    green_fill = PatternFill(start_color='aed59d', end_color='aed59d', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    sourcetarget_letter = get_column_letter(len_sourcetarget)
    start_source_letter = get_column_letter(len_sourcetarget + 1)
    end_source_letter = get_column_letter(len_source+len_sourcetarget)
    start_target_letter = get_column_letter(len_sourcetarget+len_source + 1)
    end_target_letter = get_column_letter(len_sourcetarget+len_source+len_target)

    for col in range(1, len_sourcetarget + 1):
        for cell in ws[f"A{col}":f"{sourcetarget_letter}{col}"][0]:
            cell.fill = fiolet_fill
        for cell in ws[f"{start_target_letter}{col}":f"{end_target_letter}{col}"][0]:
            cell.fill = blue_fill

        for cell in ws[f"A{col}:{end_target_letter}{col}"][0]:
            cell.border = border

    for cell in ws[f"{start_source_letter}{1}":f"{end_source_letter}{1}"][0]:
        cell.fill = orange_fill
    for cell in ws[f"{start_source_letter}{2}":f"{end_source_letter}{2}"][0]:
        cell.fill = green_fill

    ws.merge_cells(f"A1:{sourcetarget_letter}1")
    ws["A1"] = "Source/Target"

    ws.merge_cells(f"{start_source_letter}1:{end_source_letter}1")
    ws[f"{start_source_letter}1"] = "Source"

    ws.merge_cells(f"{start_target_letter}1:{end_target_letter}1")
    ws[f"{start_target_letter}1"] = "Target"

    # Выравнивание текста по центру в объединенных ячейках
    center_alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].alignment = center_alignment
    ws[f'{start_source_letter}1'].alignment = center_alignment
    ws[f'{start_target_letter}1'].alignment = center_alignment

    wb.save(result_mappings_path)


window = Tk()
window.title("Доработка маппинга")
window.geometry('600x200')

old_mappings_path = StringVar()
new_mappings_path = StringVar()
result_mappings_path = StringVar()

old_mappings_path_lbl = Label(window, text="Выберите путь старых маппингов:")
old_mappings_path_txt = Entry(window, width=50, textvariable=old_mappings_path, bg='white')
old_mappings_path_btn = Button(window, text="...", command=lambda: old_mappings_path.set(filedialog.askdirectory()), bg='white', width=2)

old_mappings_path_lbl.place(x=10, y=10) 
old_mappings_path_txt.place(x=255, y=13)
old_mappings_path_btn.place(x=555, y=8)


new_mappings_path_lbl = Label(window, text="Выберите путь новых маппингов:")
new_mappings_path_txt = Entry(window, width=50, textvariable=new_mappings_path, bg='white')
new_mappings_path_btn = Button(window, text="...", command=lambda: new_mappings_path.set(filedialog.askdirectory()), bg='white', width=2)

new_mappings_path_lbl.place(x=10, y=35) 
new_mappings_path_txt.place(x=255, y=38)
new_mappings_path_btn.place(x=555, y=33)

result_mappings_path_lbl = Label(window, text="Выберите папку сохранения результата:")
result_mappings_path_txt = Entry(window, width=50, textvariable=result_mappings_path, bg='white')
result_mappings_path_btn = Button(window, text="...", command=lambda: result_mappings_path.set(filedialog.askdirectory()), bg='white', width=2)

result_mappings_path_lbl.place(x=10, y=60) 
result_mappings_path_txt.place(x=255, y=62)
result_mappings_path_btn.place(x=555, y=58)


result_btn = Button(window, text="Старт", width=15, height=2,bg='white',
                    command=lambda: do_rework(old_mappings_path.get(),
                                         new_mappings_path.get(), 
                                         result_mappings_path.get()))
result_btn.place(x=463, y=88) 

window.mainloop()